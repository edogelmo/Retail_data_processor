#!/usr/bin/env python
# coding: utf-8

import re
import html
import warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.simplefilter("ignore", UserWarning)

# ------------------------------------------------------------
# PATHS
# ------------------------------------------------------------
BASE_DIR = Path(r"C:\Data\Stores\Performance")
MERGERS_DIR = Path(r"C:\Data\Stores\Mergers")
OUTPUT_FILE = Path(r"C:\Data\Stores\Output") / "sales_report.xlsx"

STORE_UPGRADES_FILE = Path(r"C:\Data\Stores\Upgrades\01_Upgrades_processed.xlsx")
EXPIRING_CONTRACTS_DIR = Path(r"C:\Data\Stores\Expiring_Contracts")
CUSTOMER_LOCATIONS_DIR = Path(r"C:\Data\Stores\Customer_Locations")
PREMIUM_SERVICES_DIR = Path(r"C:\Data\Stores\Premium_Services")
PROMO_INDEX_DIR = Path(r"C:\Data\Stores\Promo_Index")

FILE_PREFIX_2025 = "Performance_box_2025"
FILE_PREFIX_2024 = "Performance_box_2024"
FILE_PREFIX_MERGERS = "G-01-NET-01 - Store Network"

# ------------------------------------------------------------
# CONSTANTS & CONFIGURATION
# ------------------------------------------------------------
ADJ_ONE_TIME_DIVISOR = 10.0

def percent_points_to_ratio(x):
    s = pd.to_numeric(x, errors="coerce")
    return s / 100.0

def ratio_to_percent_points(x):
    s = pd.to_numeric(x, errors="coerce")
    return s * 100.0

PERF_REQUIRED_HEADERS = ["Store Code & Desc", "Domain Code", "Domain Desc", "Label", "Value"]
MERGERS_REQUIRED_HEADERS = ["Store Code", "Target Store Code", "Closure Date"]
UPGRADES_REQUIRED_HEADERS = ["Store Code", "Fiscal Year", "Quarter", "Upgradable Items", "Upgraded Items"]
CONTRACTS_REQUIRED_HEADERS = ["Product", "Expiring Contracts", "Renewed Contracts"]
LOCATIONS_REQUIRED_HEADERS = ["Store Code", "Customer ID", "Order Num", "Order Manager", "City", "Sales"]
PREMIUM_SVC_REQUIRED_HEADERS = ["Premium Svc", "Order"]
PROMO_INDEX_REQUIRED_CONTAINS = ["Product Aggregation", "OBSERVED PROMO INDEX"]

# Metrics from Performance file
METRIC_SALES_ELEC = "Sales Electronics"
METRIC_SALES_B2B = "Sales B2B"
METRIC_SALES_APPAREL = "Sales Apparel"
METRIC_SALES_SUBS = "Sales Subscriptions"
METRIC_SALES_ONE_TIME = "Sales One-Time"

METRIC_NC_SUBS = "NC Subscriptions" 
METRIC_NC_ONE_TIME = "NC One-Time"
METRIC_NC_APPAREL = "NC Apparel"

METRIC_PCT_ELEC = "% Mix Portfolio Electronics"
METRIC_PCT_APPAREL = "% Mix Portfolio Apparel"
METRIC_PCT_HOME = "% Mix Portfolio Home"

METRIC_PCT_AFFILIATE = "% Sales via Affiliate"
METRIC_PCT_THIRD_PARTY = "% Sales via Third Party"

METRIC_RETURNS_SALES_ELEC = "10. Returns / Sales"
METRIC_RETURNS_SALES_APPAREL = "6. Returns / Sales"

# Staff / Headcount Metrics
METRIC_NUM_ADMIN = "Number of Admins"
METRIC_NUM_LEADS = "Number of Leads"
METRIC_NUM_CLA = "Number of CLA"
METRIC_NUM_OP = "Number of OP"
METRIC_NUM_TP = "Number of Third Parties"
METRIC_NUM_CUSTOMERS = "Total Customers"

# Output Fields
FIELD_SALES_2025 = "Sales YTD"
FIELD_SALES_2025_ADJ = "Sales YTD (Merger Adj)"
FIELD_SALES_2024 = "Sales YTD-1"
FIELD_SALES_2024_ADJ = "Sales YTD-1 (Merger Adj)"
FIELD_SALES_GROWTH = "Sales Growth"

FIELD_SALES_SUBS_2025 = "Subscription Sales YTD"
FIELD_SALES_SUBS_2025_ADJ = "Subscription Sales YTD (Merger Adj)"

FIELD_PCT_SALES_ELEC = "% Sales Electronics"
FIELD_PCT_SALES_ELEC_ADJ = "% Sales Electronics (Merger Adj)"
FIELD_PCT_SALES_APP = "% Sales Apparel"
FIELD_PCT_SALES_APP_ADJ = "% Sales Apparel (Merger Adj)"
FIELD_PCT_SALES_HOME = "% Sales Home"
FIELD_PCT_SALES_HOME_ADJ = "% Sales Home (Merger Adj)"
FIELD_PCT_SALES_HOME_APP = "% Sales Home and Apparel"
FIELD_PCT_SALES_HOME_APP_ADJ = "% Sales Home and Apparel (Merger Adj)"

FIELD_PTF_ELEC = "% Portfolio Electronics"
FIELD_PTF_APP = "% Portfolio Apparel"
FIELD_PTF_HOME = "% Portfolio Home"

FIELD_PTF_AFFILIATE = "% Portfolio Affiliate"
FIELD_PCT_SALES_AFFILIATE = "% Sales Affiliate"
FIELD_PCT_SALES_DIRECT = "% Sales Direct"
FIELD_PCT_SALES_DIRECT_ONLY = "% Sales Direct (No Affiliate/Third Party)"
FIELD_RETURNS_PTF = "% Returns/Sales Portfolio"

FIELD_NC_HOME_2025 = "Home - NC YTD"
FIELD_NC_HOME_2025_ADJ = "Home - NC YTD (Merger Adj)"
FIELD_NC_HOME_2024 = "Home - NC YTD-1"
FIELD_NC_HOME_2024_ADJ = "Home - NC YTD-1 (Merger Adj)"
FIELD_VAR_NC_HOME = "Home - NC Variation"

FIELD_NC_APP_2025 = "Apparel - NC YTD"
FIELD_NC_APP_2025_ADJ = "Apparel - NC YTD (Merger Adj)"
FIELD_NC_APP_2024 = "Apparel - NC YTD-1"
FIELD_NC_APP_2024_ADJ = "Apparel - NC YTD-1 (Merger Adj)"
FIELD_VAR_NC_APP = "Apparel - NC Variation"

FIELD_NC_TOT_2025 = "Home & Apparel - NC YTD"
FIELD_NC_TOT_2025_ADJ = "Home & Apparel - NC YTD (Merger Adj)"
FIELD_NC_TOT_2024 = "Home & Apparel - NC YTD-1"
FIELD_NC_TOT_2024_ADJ = "Home & Apparel - NC YTD-1 (Merger Adj)"
FIELD_VAR_NC_TOT = "Home & Apparel - NC Variation"

FIELD_NC_SUBS_2025 = "NC Subscriptions YTD"
FIELD_NC_SUBS_2025_ADJ = "NC Subscriptions YTD (Merger Adj)"

FIELD_NC_TOT_VS_SALES = "Home & Apparel - NC/Sales"
FIELD_NC_APP_VS_SALES = "Apparel - NC/Sales"
FIELD_NC_HOME_VS_SALES = "Home - NC/Sales"

FIELD_NUM_ADMIN = "# Admins"
FIELD_ADMIN_2025_ADJ = "# Admins (Merger Adj)"
FIELD_NUM_CUSTOMERS = "# Customers"
FIELD_CUST_2025_ADJ = "# Customers (Merger Adj)"
FIELD_SALES_STAFF = "# Sales Staff"
FIELD_STAFF_2025_ADJ = "# Sales Staff (Merger Adj)"
FIELD_HEADCOUNT = "# Headcount"
FIELD_HC_2025_ADJ = "# Headcount (Merger Adj)"

FIELD_RATIO_STAFF_ADMIN = "Ratio Staff/Admins"
FIELD_RATIO_STAFF_ADMIN_ADJ = "Ratio Staff/Admins (Merger Adj)"

FIELD_SALES_PER_HC = "Sales/#Headcount"
FIELD_NC_PER_CUST = "NC/Customer"
FIELD_NC_PER_STAFF = "NC/Staff"
FIELD_NC_PER_HC = "NC/#Headcount"

FIELD_STAFF_2024 = "# Sales Staff YTD-1"
FIELD_STAFF_2024_ADJ = "# Sales Staff YTD-1 (Merger Adj)"
FIELD_GROWTH_STAFF_PCT = "% Growth Sales Staff"
FIELD_GROWTH_STAFF_ABS = "Growth Sales Staff"

FIELD_UPGRADE_RATE = "Apparel - % Upgrades"
FIELD_RENEWAL_RATE = "Home - % Renewals"
FIELD_CONC_CUST_TOP100 = "Concentration Index (Customers)"
FIELD_CONC_PROD_TOP1 = "Concentration Index (Product)"
FIELD_PREMIUM_SVC_RATE = "Home - % Premium Svc"
FIELD_PROMO_INDEX = "Electronics - Promo Index"

TIER_1_SPEC = "<3M, Elec >=60%"
TIER_2_LT3_LT60 = "<3M, Elec <60%"
TIER_3_3_9M = ">=3M and <9M"
TIER_4_GE9M = ">=9M"
TIER_5_XXL = "(XXL)"
XXL_CODES = {"I53", "I51"}

MONETARY_METRICS = {
    METRIC_SALES_ELEC, METRIC_SALES_B2B, METRIC_SALES_APPAREL, METRIC_SALES_SUBS, METRIC_SALES_ONE_TIME,
    METRIC_NC_SUBS, METRIC_NC_ONE_TIME, METRIC_NC_APPAREL,
}
MONETARY_SCALE_FACTOR = 1000.0

NON_QUANTIFIABLE_FIELDS = {
    FIELD_RENEWAL_RATE, FIELD_CONC_CUST_TOP100, FIELD_CONC_PROD_TOP1,
    FIELD_PREMIUM_SVC_RATE, FIELD_PROMO_INDEX,
}

OUTPUT_COLUMNS = [
    ("Registry", "Store Code"),
    ("Registry", "Store Name"),
    ("Registry", "Store Tier"),
    ("Registry", "Closure Date"),
    ("Registry", "Received Merger"),
    ("Registry", "Stores Absorbed"),
    ("Registry", "Merged Into"),

    ("Revenue", FIELD_SALES_2025),
    ("Revenue", FIELD_SALES_2025_ADJ),
    ("Revenue", FIELD_SALES_2024),
    ("Revenue", FIELD_SALES_2024_ADJ),
    ("Revenue", FIELD_SALES_GROWTH),

    ("Revenue", FIELD_SALES_SUBS_2025),
    ("Revenue", FIELD_SALES_SUBS_2025_ADJ),

    ("Revenue", FIELD_PCT_SALES_ELEC),
    ("Revenue", FIELD_PCT_SALES_ELEC_ADJ),
    ("Revenue", FIELD_PCT_SALES_APP),
    ("Revenue", FIELD_PCT_SALES_APP_ADJ),
    ("Revenue", FIELD_PCT_SALES_HOME),
    ("Revenue", FIELD_PCT_SALES_HOME_ADJ),
    ("Revenue", FIELD_PCT_SALES_HOME_APP),
    ("Revenue", FIELD_PCT_SALES_HOME_APP_ADJ),

    ("Revenue", FIELD_PTF_ELEC),
    ("Revenue", FIELD_PTF_APP),
    ("Revenue", FIELD_PTF_HOME),

    ("Revenue", FIELD_PTF_AFFILIATE),
    ("Revenue", FIELD_PCT_SALES_AFFILIATE),
    ("Revenue", FIELD_PCT_SALES_DIRECT),
    ("Revenue", FIELD_PCT_SALES_DIRECT_ONLY),

    ("Revenue", FIELD_RETURNS_PTF),
    ("Revenue", FIELD_CONC_CUST_TOP100),
    ("Revenue", FIELD_CONC_PROD_TOP1),

    ("Productivity", FIELD_NC_HOME_2025),
    ("Productivity", FIELD_NC_HOME_2025_ADJ),
    ("Productivity", FIELD_NC_HOME_2024),
    ("Productivity", FIELD_NC_HOME_2024_ADJ),
    ("Productivity", FIELD_VAR_NC_HOME),

    ("Productivity", FIELD_NC_APP_2025),
    ("Productivity", FIELD_NC_APP_2025_ADJ),
    ("Productivity", FIELD_NC_APP_2024),
    ("Productivity", FIELD_NC_APP_2024_ADJ),
    ("Productivity", FIELD_VAR_NC_APP),

    ("Productivity", FIELD_NC_TOT_2025),
    ("Productivity", FIELD_NC_TOT_2025_ADJ),
    ("Productivity", FIELD_NC_TOT_2024),
    ("Productivity", FIELD_NC_TOT_2024_ADJ),

    ("Productivity", FIELD_NC_SUBS_2025),
    ("Productivity", FIELD_NC_SUBS_2025_ADJ),

    ("Productivity", FIELD_NC_TOT_VS_SALES),
    ("Productivity", FIELD_NC_HOME_VS_SALES),
    ("Productivity", FIELD_NC_APP_VS_SALES),
    ("Productivity", FIELD_VAR_NC_TOT),

    ("Productivity", FIELD_NUM_ADMIN),
    ("Productivity", FIELD_ADMIN_2025_ADJ),
    ("Productivity", FIELD_SALES_STAFF),
    ("Productivity", FIELD_STAFF_2025_ADJ),

    ("Productivity", FIELD_RATIO_STAFF_ADMIN),
    ("Productivity", FIELD_RATIO_STAFF_ADMIN_ADJ),

    ("Productivity", FIELD_HEADCOUNT),
    ("Productivity", FIELD_HC_2025_ADJ),
    ("Productivity", FIELD_SALES_PER_HC),

    ("Productivity", FIELD_NUM_CUSTOMERS),
    ("Productivity", FIELD_CUST_2025_ADJ),

    ("Productivity", FIELD_NC_PER_CUST),
    ("Productivity", FIELD_NC_PER_STAFF),
    ("Productivity", FIELD_NC_PER_HC),

    ("Productivity", FIELD_UPGRADE_RATE),
    ("Productivity", FIELD_RENEWAL_RATE),
    ("Productivity", FIELD_PREMIUM_SVC_RATE),
    ("Productivity", FIELD_PROMO_INDEX),

    ("Productivity", FIELD_STAFF_2024),
    ("Productivity", FIELD_STAFF_2024_ADJ),
    ("Productivity", FIELD_GROWTH_STAFF_PCT),
    ("Productivity", FIELD_GROWTH_STAFF_ABS),
]

PERCENT_FIELDS_RATIO = {
    FIELD_SALES_GROWTH,
    FIELD_PCT_SALES_ELEC, FIELD_PCT_SALES_ELEC_ADJ,
    FIELD_PCT_SALES_APP, FIELD_PCT_SALES_APP_ADJ,
    FIELD_PCT_SALES_HOME, FIELD_PCT_SALES_HOME_ADJ,
    FIELD_PCT_SALES_HOME_APP, FIELD_PCT_SALES_HOME_APP_ADJ,
    FIELD_PTF_ELEC, FIELD_PTF_APP, FIELD_PTF_HOME,
    FIELD_RETURNS_PTF, FIELD_CONC_CUST_TOP100, FIELD_CONC_PROD_TOP1,
    FIELD_VAR_NC_HOME, FIELD_VAR_NC_APP, FIELD_VAR_NC_TOT,
    FIELD_NC_TOT_VS_SALES, FIELD_NC_HOME_VS_SALES, FIELD_NC_APP_VS_SALES,
    FIELD_GROWTH_STAFF_PCT, FIELD_UPGRADE_RATE, FIELD_RENEWAL_RATE,
    FIELD_PREMIUM_SVC_RATE, FIELD_PROMO_INDEX,
}

PERCENT_FIELDS_POINTS = {
    FIELD_PTF_AFFILIATE, FIELD_PCT_SALES_AFFILIATE,
    FIELD_PCT_SALES_DIRECT, FIELD_PCT_SALES_DIRECT_ONLY,
}

VIEW_FIELDS = [
    FIELD_SALES_GROWTH, FIELD_PCT_SALES_HOME_APP_ADJ, FIELD_PCT_SALES_APP_ADJ,
    FIELD_PCT_SALES_DIRECT, FIELD_PCT_SALES_DIRECT_ONLY, FIELD_UPGRADE_RATE,
    FIELD_RENEWAL_RATE, FIELD_PREMIUM_SVC_RATE, FIELD_PROMO_INDEX,
    FIELD_RETURNS_PTF, FIELD_CONC_CUST_TOP100, FIELD_CONC_PROD_TOP1,
    FIELD_GROWTH_STAFF_ABS, FIELD_GROWTH_STAFF_PCT, FIELD_NC_TOT_2025_ADJ,
    FIELD_NC_APP_VS_SALES, FIELD_NC_HOME_VS_SALES, FIELD_VAR_NC_TOT,
    FIELD_VAR_NC_APP, FIELD_VAR_NC_HOME, FIELD_SALES_PER_HC,
    FIELD_NC_PER_CUST, FIELD_NC_PER_STAFF, FIELD_NC_PER_HC,
]

# ------------------------------------------------------------
# CORE UTILS
# ------------------------------------------------------------
def normalize_header_text(value) -> str:
    if pd.isna(value): return ""
    text = str(value)
    for _ in range(3):
        new_text = html.unescape(text)
        if new_text == text: break
        text = new_text
    return re.sub(r"\s+", " ", text.replace("\n", " ").replace("\r", " ")).strip().lower()

def standardize_string(value):
    if pd.isna(value): return np.nan
    return re.sub(r"\s+", " ", str(value)).strip()

def parse_number(value):
    if pd.isna(value): return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)): return float(value)
    s = str(value).strip().replace("\u00A0", " ").replace(" ", "").replace("%", "")
    if s == "" or s == "-": return np.nan
    s = s.replace(".", "").replace(",", ".") if "." in s and "," in s else s.replace(",", ".")
    s = re.sub(r"[^0-9\.\-]", "", s)
    try: return float(s)
    except: return np.nan

def parse_date(value):
    return pd.to_datetime(value, dayfirst=True, errors="coerce")

def normalize_store_code(v):
    if pd.isna(v): return np.nan
    s = re.sub(r"\s+", "", str(v).upper())
    if s.isdigit(): return s.zfill(3)[-3:]
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s if s else np.nan

def safe_divide(n, d):
    n_num = pd.to_numeric(n, errors="coerce")
    d_num = pd.to_numeric(d, errors="coerce")
    
    if np.isscalar(n_num) and np.isscalar(d_num):
        if pd.isna(n_num) or pd.isna(d_num) or d_num == 0: return np.nan
        return float(n_num) / float(d_num)

    if isinstance(n_num, pd.Series) and np.isscalar(d_num):
        out = pd.Series(np.nan, index=n_num.index, dtype="float64")
        if pd.isna(d_num) or d_num == 0: return out
        ok = n_num.notna()
        out.loc[ok] = n_num.loc[ok] / float(d_num)
        return out

    if np.isscalar(n_num) and isinstance(d_num, pd.Series):
        out = pd.Series(np.nan, index=d_num.index, dtype="float64")
        if pd.isna(n_num): return out
        ok = d_num.notna() & (d_num != 0)
        out.loc[ok] = float(n_num) / d_num.loc[ok]
        return out

    out = pd.Series(np.nan, index=n_num.index, dtype="float64")
    ok = d_num.notna() & (d_num != 0) & n_num.notna()
    out.loc[ok] = n_num.loc[ok] / d_num.loc[ok]
    return out

def apply_metric_scale(series, metric_name: str):
    s = pd.to_numeric(series, errors="coerce").copy()
    if metric_name in MONETARY_METRICS:
        s = s * MONETARY_SCALE_FACTOR
    return s

def sum_series(*series_list, index=None):
    if not series_list: raise ValueError("Requires at least one series")
    if index is None: index = series_list[0].index
    df = pd.concat([pd.to_numeric(s.reindex(index), errors="coerce") for s in series_list], axis=1)
    all_nan = df.isna().all(axis=1)
    out = df.fillna(0).sum(axis=1)
    out.loc[all_nan] = np.nan
    return out

def list_excel_files(folder: Path):
    if not folder or not Path(folder).exists(): return []
    files = list(Path(folder).glob("*.xlsx")) + list(Path(folder).glob("*.xlsm")) + list(Path(folder).glob("*.xls"))
    return sorted([p for p in files if not p.name.startswith("~$")], key=lambda x: x.name.lower())

def find_excel_file(folder: Path, prefix: str) -> Path:
    matched = [p for p in list_excel_files(folder) if p.stem.lower().startswith(prefix.lower())]
    if not matched: raise FileNotFoundError(f"No file with prefix '{prefix}' in {folder}")
    return matched[0]

def read_excel_with_dynamic_header(file_path: Path, req_headers, sheet_name=0) -> pd.DataFrame:
    engine = "openpyxl" if file_path.suffix.lower() in {".xlsx", ".xlsm"} else "xlrd"
    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=object, engine=engine, nrows=180)
    req_norm = {normalize_header_text(h) for h in req_headers}
    header_row = next((i for i in range(len(raw)) if req_norm.issubset({normalize_header_text(v) for v in raw.iloc[i].tolist() if normalize_header_text(v)})), None)
    if header_row is None: raise ValueError(f"Header not found in {file_path.name}")
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, dtype=object, engine=engine)
    df.columns = [standardize_string(c) if pd.notna(c) else c for c in df.columns]
    return df.dropna(how="all").copy()

def map_series_to_master(series_by_code, codes_by_label, master_index):
    out = pd.Series(np.nan, index=master_index, dtype="float64")
    if series_by_code is None or len(series_by_code) == 0: return out
    s = series_by_code.copy()
    s.index = [normalize_store_code(x) for x in s.index]
    for lbl in master_index:
        code = normalize_store_code(codes_by_label.get(lbl, np.nan))
        if pd.notna(code) and code in s.index: out.loc[lbl] = s.loc[code]
    return out

def map_mask_to_master(mask_by_code, codes_by_label, master_index):
    out = pd.Series(False, index=master_index, dtype="bool")
    if mask_by_code is None or len(mask_by_code) == 0: return out
    m = mask_by_code.copy()
    m.index = [normalize_store_code(x) for x in m.index]
    for lbl in master_index:
        code = normalize_store_code(codes_by_label.get(lbl, np.nan))
        if pd.notna(code) and code in m.index: out.loc[lbl] = bool(m.loc[code])
    return out

# ------------------------------------------------------------
# SPECIFIC MODULES (Promo, Locations, Services, Mergers)
# ------------------------------------------------------------
def build_merger_maps(mergers_df: pd.DataFrame):
    rec_to_send, send_to_rec, send_to_cess = {}, {}, {}
    if mergers_df is None or mergers_df.empty: return rec_to_send, send_to_rec, send_to_cess
    for _, r in mergers_df.iterrows():
        s = normalize_store_code(r.get("Store Code"))
        rcv = normalize_store_code(r.get("Target Store Code"))
        cess = parse_date(r.get("Closure Date", pd.NaT))
        if pd.isna(s) or pd.isna(rcv) or s == rcv: continue
        rec_to_send.setdefault(rcv, []).append(s)
        send_to_rec[s] = rcv
        send_to_cess[s] = cess
    for k in rec_to_send: rec_to_send[k] = sorted(list(set(rec_to_send[k])))
    return rec_to_send, send_to_rec, send_to_cess

def apply_store_mergers_additive(base_2024, base_2025, codes, rec_to_send, lbl_24, lbl_25, send_to_cess):
    adj_24 = pd.to_numeric(base_2024, errors="coerce").astype("float64")
    adj_25 = pd.to_numeric(base_2025, errors="coerce").astype("float64")
    m_24, m_25 = pd.Series(False, index=adj_24.index), pd.Series(False, index=adj_25.index)

    for idx in adj_24.index:
        rcv = normalize_store_code(codes.get(idx, np.nan))
        if pd.isna(rcv) or rcv not in rec_to_send: continue
        add_24, add_25 = [], []
        for s in rec_to_send[rcv]:
            cess = send_to_cess.get(s, pd.NaT)
            if pd.isna(cess): continue
            y = pd.to_datetime(cess).year
            if y in (2025, 2026) and lbl_24.get(s) in base_2024.index:
                v = base_2024.loc[lbl_24[s]]
                if pd.notna(v): add_24.append(float(v))
            if y == 2026 and lbl_25.get(s) in base_2025.index:
                v = base_2025.loc[lbl_25[s]]
                if pd.notna(v): add_25.append(float(v))
        if add_24:
            adj_24.loc[idx] = (0.0 if pd.isna(adj_24.loc[idx]) else float(adj_24.loc[idx])) + float(np.sum(add_24))
            m_24.loc[idx] = True
        if add_25:
            adj_25.loc[idx] = (0.0 if pd.isna(adj_25.loc[idx]) else float(adj_25.loc[idx])) + float(np.sum(add_25))
            m_25.loc[idx] = True
    return adj_24, adj_25, m_24, m_25

def apply_store_mergers_average(rate_by_code, rec_to_send, send_to_cess, fusion_year=2026):
    if rate_by_code is None or len(rate_by_code) == 0: return pd.Series(dtype="float64"), pd.Series(dtype="bool")
    adj = pd.to_numeric(rate_by_code, errors="coerce").copy()
    adj.index = [normalize_store_code(x) for x in adj.index]
    
    for rcv in rec_to_send.keys():
        rcvn = normalize_store_code(rcv)
        if pd.notna(rcvn) and rcvn not in adj.index: adj.loc[rcvn] = np.nan
    mask = pd.Series(False, index=adj.index)

    for rcv in adj.index.tolist():
        if pd.isna(rcv) or rcv not in rec_to_send: continue
        values = [float(adj.get(rcv))] if pd.notna(adj.get(rcv)) else []
        added = False
        for s in rec_to_send.get(rcv, []):
            snd = normalize_store_code(s)
            cess = send_to_cess.get(snd, send_to_cess.get(s, pd.NaT))
            if pd.isna(cess) or pd.to_datetime(cess).year != fusion_year: continue
            v = adj.get(snd, np.nan)
            if pd.notna(v):
                values.append(float(v))
                added = True
        if values: adj.loc[rcv] = float(np.mean(values))
        if added: mask.loc[rcv] = True
    return adj.sort_index(), mask.reindex(adj.index, fill_value=False)

def read_promo_index(folder: Path):
    rows = []
    for fp in list_excel_files(folder):
        code = normalize_store_code(fp.stem)
        if pd.isna(code): continue
        engine = "openpyxl" if fp.suffix.lower() in {".xlsx", ".xlsm"} else "xlrd"
        try: sheets = pd.ExcelFile(fp, engine=engine).sheet_names
        except: continue
        val = np.nan
        for sh in sheets:
            try: raw = pd.read_excel(fp, sheet_name=sh, header=None, dtype=object, engine=engine, nrows=4000)
            except: continue
            norm = raw.map(normalize_header_text) if hasattr(raw, "map") else raw.apply(lambda col: col.map(normalize_header_text))
            for i in range(norm.shape[0]):
                r_vals = norm.iloc[i].tolist()
                if "product aggregation" in r_vals and any(v and "observed promo index" in v for v in r_vals):
                    col_agg = r_vals.index("product aggregation")
                    col_prx = next(j for j, v in enumerate(r_vals) if v and "observed promo index" in v)
                    data = raw.iloc[i+2:].copy()
                    agg = data.iloc[:, col_agg].astype("string").str.strip().str.upper()
                    prx = data.iloc[:, col_prx].apply(parse_number)
                    mask = agg.str.contains(r"\bTOTAL\b", na=False) & agg.str.contains(r"\bELECTRONICS\b", na=False)
                    if mask.any():
                        v_pts = prx.loc[mask].iloc[-1]
                        if pd.notna(v_pts): val = float(v_pts) / 100.0
                    break
            if pd.notna(val): break
        if pd.notna(val): rows.append({"Store Code": code, FIELD_PROMO_INDEX: val})
    if not rows: return pd.Series(dtype="float64", name=FIELD_PROMO_INDEX)
    return pd.DataFrame(rows).set_index("Store Code")[FIELD_PROMO_INDEX]

def read_premium_svc(folder: Path):
    rows = []
    for fp in list_excel_files(folder):
        code = normalize_store_code(fp.stem)
        if pd.isna(code): continue
        try: df = read_excel_with_dynamic_header(fp, PREMIUM_SVC_REQUIRED_HEADERS)
        except: continue
        c_fee = next((c for c in df.columns if "premium" in normalize_header_text(c)), None)
        if not c_fee: continue
        flags = df[c_fee].astype("string").str.strip().str.upper().replace({"<NA>": pd.NA})
        tot, mfee = int(flags.notna().sum()), int((flags == "S").sum())
        rows.append({"Store Code": code, "tot": tot, "prem": mfee})
    if not rows: return pd.DataFrame(columns=["tot", "prem"]).rename_axis("Store Code")
    return pd.DataFrame(rows).set_index("Store Code")

def read_customer_cities(folder: Path):
    rows = []
    for fp in list_excel_files(folder):
        try: df = read_excel_with_dynamic_header(fp, LOCATIONS_REQUIRED_HEADERS)
        except: continue
        cm = {col: normalize_header_text(col) for col in df.columns}
        inv = {v: k for k, v in cm.items()}
        req = ["store code", "customer id", "city", "sales"]
        if not all(k in inv for k in req): continue
        tmp = df[[inv[k] for k in req]].copy()
        tmp.columns = ["Store Code", "Customer ID", "City", "Sales"]
        tmp["Store Code"] = tmp["Store Code"].apply(normalize_store_code)
        tmp["Customer ID"] = tmp["Customer ID"].apply(standardize_string)
        tmp["City"] = tmp["City"].apply(standardize_string)
        tmp["Sales"] = tmp["Sales"].apply(parse_number)
        tmp = tmp.dropna(subset=["Store Code", "Sales"])
        tmp = tmp[pd.to_numeric(tmp["Sales"], errors="coerce") > 0]
        rows.append(tmp.drop_duplicates())
    if not rows: return pd.DataFrame(columns=["Store Code", "Customer ID", "City", "Sales"])
    return pd.concat(rows, ignore_index=True)

def compute_city_tables(df: pd.DataFrame, top_n=20):
    out = {}
    if df is None or df.empty: return out
    df = df.dropna(subset=["Store Code", "Customer ID", "City"])
    for code, g in df.groupby("Store Code"):
        city_counts = g.groupby("City")["Customer ID"].nunique().sort_values(ascending=False)
        tot = int(city_counts.sum())
        top = city_counts.head(top_n).reset_index()
        top.columns = ["City", "Number of Customers"]
        if tot > 0:
            top["% on total"] = safe_divide(top["Number of Customers"], tot)
            top["% cumulative"] = top["% on total"].cumsum()
        else:
            top["% on total"], top["% cumulative"] = np.nan, np.nan
        out[code] = {"customers": top}
    return out

def compute_concentration(df: pd.DataFrame, top_n=100):
    if df is None or df.empty: return pd.DataFrame(columns=[FIELD_CONC_CUST_TOP100, FIELD_CONC_PROD_TOP1])
    df = df.dropna(subset=["Store Code", "Sales"])
    df = df[df["Sales"] > 0]
    tot_st = df.groupby("Store Code")["Sales"].sum(min_count=1)
    
    df_c = df.dropna(subset=["Customer ID"])
    sc = df_c.groupby(["Store Code", "Customer ID"])["Sales"].sum(min_count=1)
    t100 = {c: safe_divide(float(s.droplevel(0).sort_values(ascending=False).head(top_n).sum()) if len(s.droplevel(0)) else np.nan, tot_st.get(c, np.nan)) for c, s in sc.groupby(level=0)}
    
    return pd.DataFrame({FIELD_CONC_CUST_TOP100: pd.Series(t100, dtype="float64")}).rename_axis("Store Code")

def compute_upgrades(file_path: Path):
    try: df = read_excel_with_dynamic_header(file_path, UPGRADES_REQUIRED_HEADERS)
    except: return pd.DataFrame()
    cm = {col: normalize_header_text(col) for col in df.columns}
    inv = {v: k for k, v in cm.items()}
    out = df[[inv["store code"], inv["fiscal year"], inv["upgradable items"], inv["upgraded items"]]].copy()
    out.columns = ["Store Code", "Year", "Upgradable", "Upgraded"]
    out["Store Code"] = out["Store Code"].apply(normalize_store_code)
    out["Year"] = pd.to_numeric(out["Year"], errors="coerce")
    out["Upgradable"] = out["Upgradable"].apply(parse_number)
    out["Upgraded"] = out["Upgraded"].apply(parse_number)
    return out.dropna(subset=["Store Code", "Year"])

def compute_expiring(folder: Path):
    rows = []
    for fp in list_excel_files(folder):
        code = normalize_store_code(fp.stem)
        if pd.isna(code): continue
        try: df = read_excel_with_dynamic_header(fp, CONTRACTS_REQUIRED_HEADERS)
        except: continue
        cm = {col: normalize_header_text(col) for col in df.columns}
        inv = {v: k for k, v in cm.items()}
        tmp = df[[inv["product"], inv["expiring contracts"], inv["renewed contracts"]]].copy()
        tmp.columns = ["Product", "Expiring", "Renewed"]
        mask = tmp["Product"].astype("string").str.strip().str.lower().eq("total")
        if mask.any():
            tot = tmp.loc[mask].iloc[-1]
            rows.append({"Store Code": code, "exp": parse_number(tot["Expiring"]), "ren": parse_number(tot["Renewed"])})
    if not rows: return pd.DataFrame(columns=["exp", "ren"]).rename_axis("Store Code")
    return pd.DataFrame(rows).set_index("Store Code").groupby(level=0).sum(numeric_only=True)

# ------------------------------------------------------------
# BUILD FINAL DATASET
# ------------------------------------------------------------
def get_perf_metric(pvt, metric, idx):
    if pvt is None or pvt.empty or metric not in pvt.columns: return pd.Series(np.nan, index=idx, dtype="float64")
    return apply_metric_scale(pd.to_numeric(pvt[metric].reindex(idx), errors="coerce"), metric)

def build_final_dataset(pvt_25, pvt_24, mergers, upgrades, contracts, conc_df, mask_conc, mfee_ratio, mask_mfee, promo, mask_promo, fusion_year=2026):
    lbl_25 = list(pvt_25.index) if pvt_25 is not None and not pvt_25.empty else []
    lbl_24 = list(pvt_24.index) if pvt_24 is not None and not pvt_24.empty else []
    idx = pd.Index(sorted(pd.unique(lbl_25 + lbl_24)))
    
    s = pd.Series(idx, index=idx).astype("string").str.strip()
    codes_by_label = s.str.extract(r"^\s*([^-]+?)\s*-\s*.*$")[0].apply(normalize_store_code)
    names_by_label = s.str.extract(r"^\s*[^-]+?\s*-\s*(.*)$")[0]
    
    c2l_25 = {normalize_store_code(c): l for l, c in zip(pvt_25.index, pvt_25.index.str.split('-').str[0])} if pvt_25 is not None else {}
    c2l_24 = {normalize_store_code(c): l for l, c in zip(pvt_24.index, pvt_24.index.str.split('-').str[0])} if pvt_24 is not None else {}
    
    rec_to_send, send_to_rec, send_to_cess = build_merger_maps(mergers)
    
    # Base extraction
    s_elec_25 = get_perf_metric(pvt_25, METRIC_SALES_ELEC, idx)
    s_b2b_25  = get_perf_metric(pvt_25, METRIC_SALES_B2B, idx)
    s_app_25  = get_perf_metric(pvt_25, METRIC_SALES_APPAREL, idx)
    s_sub_25  = get_perf_metric(pvt_25, METRIC_SALES_SUBS, idx)
    s_ot_25   = get_perf_metric(pvt_25, METRIC_SALES_ONE_TIME, idx)

    s_elec_24 = get_perf_metric(pvt_24, METRIC_SALES_ELEC, idx)
    s_b2b_24  = get_perf_metric(pvt_24, METRIC_SALES_B2B, idx)
    s_app_24  = get_perf_metric(pvt_24, METRIC_SALES_APPAREL, idx)
    s_sub_24  = get_perf_metric(pvt_24, METRIC_SALES_SUBS, idx)
    s_ot_24   = get_perf_metric(pvt_24, METRIC_SALES_ONE_TIME, idx)

    s_home_subs_25 = sum_series(s_sub_25, s_b2b_25, index=idx)
    s_home_subs_24 = sum_series(s_sub_24, s_b2b_24, index=idx)
    s_home_adj_25 = sum_series(s_home_subs_25, pd.to_numeric(s_ot_25, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)
    s_home_adj_24 = sum_series(s_home_subs_24, pd.to_numeric(s_ot_24, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)

    tot_25 = sum_series(s_elec_25, s_app_25, s_home_adj_25, index=idx)
    tot_24 = sum_series(s_elec_24, s_app_24, s_home_adj_24, index=idx)

    # Mergers applied to sales
    s_elec_24_a, s_elec_25_a, m_elec_24, m_elec_25 = apply_store_mergers_additive(s_elec_24, s_elec_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    s_app_24_a, s_app_25_a, m_app_24, m_app_25 = apply_store_mergers_additive(s_app_24, s_app_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    s_b2b_24_a, s_b2b_25_a, m_b2b_24, m_b2b_25 = apply_store_mergers_additive(s_b2b_24, s_b2b_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    s_sub_24_a, s_sub_25_a, m_sub_24, m_sub_25 = apply_store_mergers_additive(s_sub_24, s_sub_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    s_ot_24_a, s_ot_25_a, m_ot_24, m_ot_25 = apply_store_mergers_additive(s_ot_24, s_ot_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)

    s_home_subs_24_a = sum_series(s_sub_24_a, s_b2b_24_a, index=idx)
    s_home_subs_25_a = sum_series(s_sub_25_a, s_b2b_25_a, index=idx)
    s_home_adj_24_a = sum_series(s_home_subs_24_a, pd.to_numeric(s_ot_24_a, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)
    s_home_adj_25_a = sum_series(s_home_subs_25_a, pd.to_numeric(s_ot_25_a, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)

    tot_24_a = sum_series(s_elec_24_a, s_app_24_a, s_home_adj_24_a, index=idx)
    tot_25_a = sum_series(s_elec_25_a, s_app_25_a, s_home_adj_25_a, index=idx)

    # NP extraction and adjustments
    np_sub_25 = get_perf_metric(pvt_25, METRIC_NC_SUBS, idx)
    np_ot_25  = get_perf_metric(pvt_25, METRIC_NC_ONE_TIME, idx)
    np_app_25 = get_perf_metric(pvt_25, METRIC_NC_APPAREL, idx)
    np_sub_24 = get_perf_metric(pvt_24, METRIC_NC_SUBS, idx)
    np_ot_24  = get_perf_metric(pvt_24, METRIC_NC_ONE_TIME, idx)
    np_app_24 = get_perf_metric(pvt_24, METRIC_NC_APPAREL, idx)

    np_home_25 = sum_series(np_sub_25, pd.to_numeric(np_ot_25, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)
    np_home_24 = sum_series(np_sub_24, pd.to_numeric(np_ot_24, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)
    np_tot_25 = sum_series(np_app_25, np_home_25, index=idx)
    np_tot_24 = sum_series(np_app_24, np_home_24, index=idx)

    np_sub_24_a, np_sub_25_a, m_nps_24, m_nps_25 = apply_store_mergers_additive(np_sub_24, np_sub_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    np_ot_24_a, np_ot_25_a, m_npo_24, m_npo_25 = apply_store_mergers_additive(np_ot_24, np_ot_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    np_app_24_a, np_app_25_a, m_npa_24, m_npa_25 = apply_store_mergers_additive(np_app_24, np_app_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)

    np_home_24_a = sum_series(np_sub_24_a, pd.to_numeric(np_ot_24_a, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)
    np_home_25_a = sum_series(np_sub_25_a, pd.to_numeric(np_ot_25_a, errors="coerce")/ADJ_ONE_TIME_DIVISOR, index=idx)
    np_tot_24_a = sum_series(np_app_24_a, np_home_24_a, index=idx)
    np_tot_25_a = sum_series(np_app_25_a, np_home_25_a, index=idx)

    # Staff metrics
    def get_staff(pvt):
        a = get_perf_metric(pvt, METRIC_NUM_ADMIN, idx)
        l = get_perf_metric(pvt, METRIC_NUM_LEADS, idx)
        c = get_perf_metric(pvt, METRIC_NUM_CLA, idx)
        o = get_perf_metric(pvt, METRIC_NUM_OP, idx)
        t = get_perf_metric(pvt, METRIC_NUM_TP, idx)
        cust = get_perf_metric(pvt, METRIC_NUM_CUSTOMERS, idx)
        return a, sum_series(l, c, o, t, index=idx), cust
    
    adm_25, staf_25, cust_25 = get_staff(pvt_25)
    adm_24, staf_24, cust_24 = get_staff(pvt_24)

    adm_24_a, adm_25_a, m_adm_24, m_adm_25 = apply_store_mergers_additive(adm_24, adm_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    staf_24_a, staf_25_a, m_staf_24, m_staf_25 = apply_store_mergers_additive(staf_24, staf_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)
    cust_24_a, cust_25_a, m_cust_24, m_cust_25 = apply_store_mergers_additive(cust_24, cust_25, codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess)

    hc_25 = sum_series(staf_25, adm_25, index=idx)
    hc_25_a = sum_series(staf_25_a, adm_25_a, index=idx)

    # Portfolios & Ratios
    ptf_elec = percent_points_to_ratio(get_perf_metric(pvt_25, METRIC_PCT_ELEC, idx))
    ptf_app = percent_points_to_ratio(get_perf_metric(pvt_25, METRIC_PCT_APPAREL, idx))
    ptf_home = percent_points_to_ratio(get_perf_metric(pvt_25, METRIC_PCT_HOME, idx))
    
    aff_pts = pd.to_numeric(get_perf_metric(pvt_25, METRIC_PCT_AFFILIATE, idx), errors="coerce")
    tp_pts = pd.to_numeric(get_perf_metric(pvt_25, METRIC_PCT_THIRD_PARTY, idx), errors="coerce")
    dir_pts = (1 - percent_points_to_ratio(aff_pts)) * 100.0
    dir_only_pts = (100.0 - aff_pts - tp_pts).clip(lower=0.0, upper=100.0)

    ret_elec = percent_points_to_ratio(get_perf_metric(pvt_25, METRIC_RETURNS_SALES_ELEC, idx))
    ret_app = percent_points_to_ratio(get_perf_metric(pvt_25, METRIC_RETURNS_SALES_APPAREL, idx))
    ret_ptf = safe_divide((ret_elec * s_elec_25_a) + (ret_app * s_app_25_a), sum_series(s_elec_25_a, s_app_25_a, index=idx))

    # Tiers
    tiers = pd.Series(pd.NA, index=idx, dtype="object")
    is_xxl = codes_by_label.isin(XXL_CODES)
    tiers.loc[is_xxl] = TIER_5_XXL
    norm = ~is_xxl
    lt3 = tot_25_a.lt(3_000_000)
    ge9 = tot_25_a.ge(9_000_000)
    pct_e = safe_divide(s_elec_25_a, tot_25_a)
    tiers.loc[norm & lt3 & pct_e.ge(0.60)] = TIER_1_SPEC
    tiers.loc[norm & lt3 & (~pct_e.ge(0.60))] = TIER_2_LT3_LT60
    tiers.loc[norm & tot_25_a.ge(3_000_000) & tot_25_a.lt(9_000_000)] = TIER_3_3_9M
    tiers.loc[norm & ge9] = TIER_4_GE9M

    # Advanced merging (Rates)
    upg_rate = compute_upgrade_rate(upgrades)
    upg_rate_a, upg_mask = apply_store_mergers_average(upg_rate, rec_to_send, send_to_cess, fusion_year)
    upg_lbl = map_series_to_master(upg_rate_a, codes_by_label, idx)

    cap_exp_ren, cap_mask = apply_store_mergers_additive(contracts["exp"] if "exp" in contracts else pd.Series(), contracts["ren"] if "ren" in contracts else pd.Series(), codes_by_label, rec_to_send, c2l_24, c2l_25, send_to_cess) # simplified cap logic
    ren_rate = safe_divide(cap_exp_ren, cap_mask) if not cap_exp_ren.empty else pd.Series()
    ren_lbl = map_series_to_master(ren_rate, codes_by_label, idx)

    # DataFrame Building
    final = pd.DataFrame(index=idx)
    final["Store Code"] = codes_by_label
    final["Store Name"] = names_by_label
    final["Store Tier"] = tiers
    final["Closure Date"] = map_series_to_master(pd.Series({k: v for k, v in send_to_cess.items()}), codes_by_label, idx)
    final["Received Merger"] = np.where(codes_by_label.isin(rec_to_send.keys()), "Yes", "No")
    final["Stores Absorbed"] = codes_by_label.map(lambda x: ", ".join(rec_to_send.get(x, []))).replace({"": np.nan})
    final["Merged Into"] = codes_by_label.map(send_to_rec).replace({"": np.nan})

    final[FIELD_SALES_2025] = tot_25
    final[FIELD_SALES_2025_ADJ] = tot_25_a
    final[FIELD_SALES_2024] = tot_24
    final[FIELD_SALES_2024_ADJ] = tot_24_a
    final[FIELD_SALES_GROWTH] = safe_divide(tot_25_a - tot_24_a, tot_24_a)
    final[FIELD_SALES_SUBS_2025] = s_home_subs_25
    final[FIELD_SALES_SUBS_2025_ADJ] = s_home_subs_25_a

    final[FIELD_PCT_SALES_ELEC] = safe_divide(s_elec_25, tot_25)
    final[FIELD_PCT_SALES_ELEC_ADJ] = safe_divide(s_elec_25_a, tot_25_a)
    final[FIELD_PCT_SALES_APP] = safe_divide(s_app_25, tot_25)
    final[FIELD_PCT_SALES_APP_ADJ] = safe_divide(s_app_25_a, tot_25_a)
    final[FIELD_PCT_SALES_HOME] = safe_divide(s_home_adj_25, tot_25)
    final[FIELD_PCT_SALES_HOME_ADJ] = safe_divide(s_home_adj_25_a, tot_25_a)
    final[FIELD_PCT_SALES_HOME_APP] = safe_divide(sum_series(s_home_adj_25, s_app_25, index=idx), tot_25)
    final[FIELD_PCT_SALES_HOME_APP_ADJ] = safe_divide(sum_series(s_home_adj_25_a, s_app_25_a, index=idx), tot_25_a)

    final[FIELD_PTF_ELEC] = ptf_elec
    final[FIELD_PTF_APP] = ptf_app
    final[FIELD_PTF_HOME] = ptf_home
    final[FIELD_PTF_AFFILIATE] = aff_pts
    final[FIELD_PCT_SALES_AFFILIATE] = aff_pts
    final[FIELD_PCT_SALES_DIRECT] = dir_pts
    final[FIELD_PCT_SALES_DIRECT_ONLY] = dir_only_pts
    final[FIELD_RETURNS_PTF] = ret_ptf

    final[FIELD_CONC_CUST_TOP100] = map_series_to_master(conc_df.get(FIELD_CONC_CUST_TOP100) if not conc_df.empty else None, codes_by_label, idx)
    final[FIELD_CONC_PROD_TOP1] = map_series_to_master(conc_df.get(FIELD_CONC_PROD_TOP1) if not conc_df.empty else None, codes_by_label, idx)

    final[FIELD_NC_HOME_2025] = np_home_25
    final[FIELD_NC_HOME_2025_ADJ] = np_home_25_a
    final[FIELD_NC_HOME_2024] = np_home_24
    final[FIELD_NC_HOME_2024_ADJ] = np_home_24_a
    final[FIELD_VAR_NC_HOME] = safe_divide(np_home_25_a - np_home_24_a, np_home_24_a)

    final[FIELD_NC_APP_2025] = np_app_25
    final[FIELD_NC_APP_2025_ADJ] = np_app_25_a
    final[FIELD_NC_APP_2024] = np_app_24
    final[FIELD_NC_APP_2024_ADJ] = np_app_24_a
    final[FIELD_VAR_NC_APP] = safe_divide(np_app_25_a - np_app_24_a, np_app_24_a)

    final[FIELD_NC_TOT_2025] = np_tot_25
    final[FIELD_NC_TOT_2025_ADJ] = np_tot_25_a
    final[FIELD_NC_TOT_2024] = np_tot_24
    final[FIELD_NC_TOT_2024_ADJ] = np_tot_24_a
    final[FIELD_VAR_NC_TOT] = safe_divide(np_tot_25_a - np_tot_24_a, np_tot_24_a)

    final[FIELD_NC_SUBS_2025] = np_sub_25
    final[FIELD_NC_SUBS_2025_ADJ] = np_sub_25_a

    final[FIELD_NC_TOT_VS_SALES] = safe_divide(np_tot_25_a, sum_series(s_home_adj_25_a, s_app_25_a, index=idx))
    final[FIELD_NC_HOME_VS_SALES] = safe_divide(np_home_25_a, s_home_adj_25_a)
    final[FIELD_NC_APP_VS_SALES] = safe_divide(np_app_25_a, s_app_25_a)

    final[FIELD_NUM_ADMIN] = adm_25
    final[FIELD_ADMIN_2025_ADJ] = adm_25_a
    final[FIELD_SALES_STAFF] = staf_25
    final[FIELD_STAFF_2025_ADJ] = staf_25_a
    final[FIELD_RATIO_STAFF_ADMIN] = safe_divide(staf_25, adm_25)
    final[FIELD_RATIO_STAFF_ADMIN_ADJ] = safe_divide(staf_25_a, adm_25_a)

    final[FIELD_HEADCOUNT] = hc_25
    final[FIELD_HC_2025_ADJ] = hc_25_a
    final[FIELD_SALES_PER_HC] = safe_divide(tot_25_a, hc_25_a)

    final[FIELD_NUM_CUSTOMERS] = cust_25
    final[FIELD_CUST_2025_ADJ] = cust_25_a
    final[FIELD_NC_PER_CUST] = safe_divide(np_tot_25_a, cust_25_a)
    final[FIELD_NC_PER_STAFF] = safe_divide(np_tot_25_a, staf_25_a)
    final[FIELD_NC_PER_HC] = safe_divide(np_tot_25_a, hc_25_a)

    final[FIELD_UPGRADE_RATE] = upg_lbl
    final[FIELD_RENEWAL_RATE] = ren_lbl
    final[FIELD_PREMIUM_SVC_RATE] = map_series_to_master(mfee_ratio, codes_by_label, idx)
    final[FIELD_PROMO_INDEX] = map_series_to_master(promo, codes_by_label, idx)

    final[FIELD_STAFF_2024] = staf_24
    final[FIELD_STAFF_2024_ADJ] = staf_24_a
    final[FIELD_GROWTH_STAFF_PCT] = safe_divide(staf_25_a - staf_24_a, staf_24_a)
    final[FIELD_GROWTH_STAFF_ABS] = staf_25_a - staf_24_a

    # Build red masks
    rm = {}
    m_inc_25 = combine_masks(m_elec_25, m_app_25, m_b2b_25, m_sub_25, m_ot_25)
    m_inc_24 = combine_masks(m_elec_24, m_app_24, m_b2b_24, m_sub_24, m_ot_24)
    rm[FIELD_SALES_2025_ADJ] = m_inc_25
    rm[FIELD_SALES_2024_ADJ] = m_inc_24
    rm[FIELD_SALES_GROWTH] = combine_masks(m_inc_25, m_inc_24)
    
    return final, rm

# ------------------------------------------------------------
# EXCEL WRITING & FORMATTING
# ------------------------------------------------------------
def _is_percent_field(fname): return fname in PERCENT_FIELDS_RATIO or fname in PERCENT_FIELDS_POINTS or "%" in fname
def _is_monetary_field(fname): return "Sales" in fname and not _is_percent_field(fname)

def format_excel_value(fname, v):
    if pd.isna(v) or isinstance(v, (bool, np.bool_)): return v
    if isinstance(v, (int, float, np.integer, np.floating)):
        if _is_monetary_field(fname): v = float(v) / 1000.0
        if not _is_percent_field(fname):
            try: v = int(round(float(v), 0))
            except: pass
    return v

def autosize_cols(ws):
    for col in ws.columns:
        max_len = max([len(str(c.value)) for c in col if c.value] + [0])
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 2, 55))

def write_sheet(writer, df, cols, title, red_masks=None):
    ws = writer.book.create_sheet(title=title)
    bold, center, medium, thin = Font(bold=True), Alignment(horizontal="center"), Border(left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium")), Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    for i, cname in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=cname)
        c.font, c.alignment, c.border = bold, center, medium

    for r_idx, idx in enumerate(df.index, 2):
        for c_idx, cname in enumerate(cols, 1):
            v = format_excel_value(cname, df.at[idx, cname] if cname in df.columns else np.nan)
            c = ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(v) else v)
            c.border = thin
            if pd.notna(v):
                if cname in PERCENT_FIELDS_RATIO: c.number_format = "0.00%"
                elif cname in PERCENT_FIELDS_POINTS: c.number_format = '0.00"%"'
                elif isinstance(v, (int, float)): c.number_format = "#,##0"
            if red_masks and cname in red_masks and red_masks[cname].reindex(df.index).fillna(False).loc[idx]:
                c.font = Font(color="FF0000")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    autosize_cols(ws)

def build_quintiles(df):
    if df is None or df.empty: return pd.DataFrame()
    num_cols = [c for c in df.columns if c not in ["Store Code", "Store Name", "Store Tier", "Closure Date", "Received Merger", "Stores Absorbed", "Merged Into"]]
    groups = [("Total", df)] + [(t, df[df["Store Tier"] == t]) for t in [TIER_1_SPEC, TIER_2_LT3_LT60, TIER_3_3_9M, TIER_4_GE9M, TIER_5_XXL] if "Store Tier" in df.columns]
    
    rows = []
    for f in num_cols:
        row = {"Campo": f}
        for gname, gdf in groups:
            if f in NON_QUANTIFIABLE_FIELDS:
                for q in ["P20", "P40", "Median", "P60", "P80"]: row[f"{q} {gname}"] = "Not Calculable"
            else:
                s = pd.to_numeric(gdf[f], errors="coerce").dropna()
                row[f"P20 {gname}"] = s.quantile(0.2) if not s.empty else np.nan
                row[f"P40 {gname}"] = s.quantile(0.4) if not s.empty else np.nan
                row[f"Median {gname}"] = s.quantile(0.5) if not s.empty else np.nan
                row[f"P60 {gname}"] = s.quantile(0.6) if not s.empty else np.nan
                row[f"P80 {gname}"] = s.quantile(0.8) if not s.empty else np.nan
        rows.append(row)
    return pd.DataFrame(rows)

def write_excel_reports(final_df, red_masks, city_tables, output_path):
    q_df = build_quintiles(final_df)
    v_cols = ["Store Code", "Store Name", "Store Tier"] + [c for c in VIEW_FIELDS if c in final_df.columns]
    v_df = final_df[v_cols]
    vq_df = build_quintiles(v_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_sheet(writer, final_df, [c for _, c in OUTPUT_COLUMNS], "Sales Report", red_masks)
        write_sheet(writer, q_df, q_df.columns, "Sales Quintiles")
        
        # City tables
        ws_c = writer.book.create_sheet(title="Customer Cities")
        r = 1
        for code, data in sorted(city_tables.items()):
            df_c = data.get("customers", pd.DataFrame())
            ws_c.cell(row=r, column=1, value=f"Store {code}").font = Font(bold=True, size=12)
            r += 2
            if not df_c.empty:
                for j, h in enumerate(df_c.columns, 1): ws_c.cell(row=r, column=j, value=h).font = Font(bold=True)
                r += 1
                for _, row in df_c.iterrows():
                    for j, h in enumerate(df_c.columns, 1):
                        v = row[h]
                        c = ws_c.cell(row=r, column=j, value=None if pd.isna(v) else v)
                        if "%" in h: c.number_format = "0.00%"
                    r += 1
            r += 2

        write_sheet(writer, v_df, v_cols, "Summary View", {k: v for k, v in red_masks.items() if k in v_df.columns})
        write_sheet(writer, vq_df, vq_df.columns, "Summary Quintiles")
        
        for s in list(writer.book.sheetnames):
            if s not in ["Sales Report", "Sales Quintiles", "Customer Cities", "Summary View", "Summary Quintiles"]: writer.book.remove(writer.book[s])
        writer.book.active = 0

# ------------------------------------------------------------
# MAIN EXECUTION
# ------------------------------------------------------------
if __name__ == "__main__":
    try:
        pvt_25 = pivot_perf_data(read_performance_file(find_excel_file(BASE_DIR, FILE_PREFIX_2025)))
        pvt_24 = pivot_perf_data(read_performance_file(find_excel_file(BASE_DIR, FILE_PREFIX_2024)))
        merg = read_excel_with_dynamic_header(find_excel_file(MERGERS_DIR, FILE_PREFIX_MERGERS), MERGERS_REQUIRED_HEADERS)
        
        rec_to_send, send_to_rec, send_to_cess = build_merger_maps(merg)
        
        upg = compute_upgrades(STORE_UPGRADES_FILE)
        cont = compute_expiring(EXPIRING_CONTRACTS_DIR)
        
        c_long = read_customer_cities(CUSTOMER_LOCATIONS_DIR)
        c_tabs = compute_city_tables(c_long)
        c_conc = compute_concentration(c_long)
        
        mfee_r, mfee_m = apply_mergers_premium_svc(read_premium_svc(PREMIUM_SERVICES_DIR), rec_to_send, send_to_cess)
        promo_r, promo_m = apply_store_mergers_average(read_promo_index(PROMO_INDEX_DIR), rec_to_send, send_to_cess)
        
        final_df, red_masks = build_final_dataset(
            pvt_25, pvt_24, merg, upg, cont, c_conc, pd.Series(), 
            mfee_r, mfee_m, promo_r, promo_m, 2026
        )
        
        write_excel_reports(final_df, red_masks, c_tabs, OUTPUT_FILE)
        print(f"Report generated successfully: {OUTPUT_FILE}")
    except Exception as e:
        print(f"Execution Error: {e}")
